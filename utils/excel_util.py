#!/usr/bin/env python3
# -*- coding: utf-8 -*-
'''
    导出表格工具类
'''
import openpyxl
from commons.color import Colors
from commons.my_thread import MyThread

class ExcelUtils:
    
    @classmethod
    def save(self, path:str, sheet:str, data):
        '''
            保存为Excel表格
            Args:
                path:表格生成路径
                sheet:要创建的sheet页名称
                data:要导出的数据对象,示例数据：
                [\n
                    ['title1','title2','title3',……],\n
                    ['data1' ,'data2' ,'data3' ,……],\n
                    …………\n
                ]
            Returns:
                path:
        '''
        workBook = openpyxl.Workbook()
        sheetTable = workBook.create_sheet(sheet)
        for rowIndex in range(len(data)):
            print("导出Excel：正在处理 %s / %s 条数据"%(rowIndex+1, len(data)))
            for columnIndex in range(len(data[rowIndex])) :
                sheetTable.cell(row=rowIndex+1, column=columnIndex+1, value=data[rowIndex][columnIndex])
        workBook.save(path)
        print("成功导出Excel数据，路径：%s，sheet页为：%s"%(path, sheet))

    @classmethod
    def read(self, path:str, onlyTitle):
        '''
        读取Excel并转换为对象
        默认读取第一个sheet页内容并将第一行作为key
        '''
        workBook=openpyxl.load_workbook(path)
        allSheel=workBook.get_sheet_names()
        execSheel=workBook.get_sheet_by_name(allSheel[0])
        data = []
        dataTitle = []
        for row in range(1, execSheel.max_row + 1):
            dataItem = {}
            for column in range(1, execSheel.max_column + 1):
                if row == 1:
                    dataTitle.append(execSheel.cell(row,column).value)
                elif row > 1 and onlyTitle:
                    return dataTitle
                else:
                    dataItem[dataTitle[column - 1]] = execSheel.cell(row,column).value
            if len(dataItem.keys()) > 0:
                data.append(dataItem)
        
        return data

    @classmethod
    def read(self, path:str, startRow:int, readColumnIndexs, dataTitle):
        '''
        读取Excel并转换为对象
        path：Excel路径
        startRow：开始读取的行数，最小值为1
        readColumnIndexs：需要读取的列索引数组，第一列索引为1
        keyArr：读取的列索引对应的key值数组
        '''
        Colors.print(Colors.OKBLUE,"正在读取 %s 表格内容数据……"%path)
        workBook=openpyxl.load_workbook(path)
        allSheel=workBook.get_sheet_names()
        excelSheel=workBook.get_sheet_by_name(allSheel[0])
        data = []
        excelMaxRow = excelSheel.max_row + 1
        threadRange = 500
        threadCount = int((excelMaxRow/threadRange) + 1) if (excelMaxRow % threadRange) > 0 else int(excelMaxRow/threadRange)
        threadList = []
        for thread in range(0, threadCount) :
            startReadRow = startRow if thread == 0 else thread*threadRange
            endReadRow = excelMaxRow if (thread*threadRange + threadRange) > excelMaxRow else thread*threadRange + threadRange
            threadList.append(MyThread(func=self.read2, args=(path, excelSheel, startReadRow, endReadRow, readColumnIndexs, dataTitle)))
        for threadItem in threadList:
            threadItem.setDaemon(True)
            threadItem.start()
        for threadItem in threadList:
            threadItem.join()
        for threadItem in threadList:
            data.append(threadItem.getResult())
        newData = []
        for dataThreadItem in data:
            newData = newData + dataThreadItem
        return newData
    
    @classmethod
    def read2(self, path:str, excelSheel, startRow:int, endRow:int, readColumnIndexs, dataTitle):
        data = []
        for row in range(startRow, endRow):
            Colors.print(Colors.OKBLUE,"正在读取 %s 表格内容 %s / %s 条数据……"%(path, row, endRow))
            dataItem = {}
            dataTitleIndex = 0
            for column in readColumnIndexs:
                dataItem[dataTitle[dataTitleIndex]] = excelSheel.cell(row,column).value
                dataTitleIndex += 1
            if len(dataItem.keys()) > 0:
                data.append(dataItem)
        return data